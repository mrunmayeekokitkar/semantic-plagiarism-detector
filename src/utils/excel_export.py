"""
src/utils/excel_export.py
-------------------------
Utility for exporting similarity matrices into styled Excel (.xlsx) workbooks
with conditional formatting matching the application's heatmap logic.
"""

import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill


def export_similarity_matrix_to_excel(
    df: pd.DataFrame, threshold: float = 0.59
) -> bytes:
    """Exports a similarity matrix DataFrame into an Excel file (.xlsx) with formatting."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Similarity Matrix"

    # Write headers and index labels
    ws.cell(row=1, column=1, value="Document")
    for col_idx, col_name in enumerate(df.columns, start=2):
        ws.cell(row=1, column=col_idx, value=col_name)

    for row_idx, (index_label, row) in enumerate(df.iterrows(), start=2):
        ws.cell(row=row_idx, column=1, value=index_label)
        for col_idx, val in enumerate(row, start=2):
            cell = ws.cell(row=row_idx, column=col_idx, value=float(val))
            cell.number_format = "0.0%"
            cell.alignment = Alignment(horizontal="right")

    # Header styling
    header_fill = PatternFill(
        start_color="1F2937", end_color="1F2937", fill_type="solid"
    )
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font

    # Apply Conditional Formatting (3-Color Scale)
    max_row = len(df) + 1
    max_col = len(df.columns) + 1

    if max_row > 1 and max_col > 1:
        start_cell = "B2"
        end_col_letter = ws.cell(row=max_row, column=max_col).column_letter
        end_cell = f"{end_col_letter}{max_row}"
        matrix_range = f"{start_cell}:{end_cell}"

        color_scale = ColorScaleRule(
            start_type="num",
            start_value=0.0,
            start_color="FFFFFF",  # White (0%)
            mid_type="num",
            mid_value=threshold,
            mid_color="FEF08A",  # Yellow (At threshold)
            end_type="num",
            end_value=1.0,
            end_color="EF4444",  # Red (100%)
        )
        ws.conditional_formatting.add(matrix_range, color_scale)

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
