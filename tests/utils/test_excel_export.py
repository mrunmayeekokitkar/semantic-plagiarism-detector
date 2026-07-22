"""
tests/utils/test_excel_export.py
---------------------------------
Unit tests for styled Excel export utility.
"""

import io

import pandas as pd
from openpyxl import load_workbook

from src.utils.excel_export import export_similarity_matrix_to_excel


def test_export_similarity_matrix_to_excel():
    data = {
        "doc1.pdf": [1.0, 0.95, 0.20],
        "doc2.pdf": [0.95, 1.0, 0.15],
        "doc3.pdf": [0.20, 0.15, 1.0],
    }
    df = pd.DataFrame(data, index=["doc1.pdf", "doc2.pdf", "doc3.pdf"])

    excel_bytes = export_similarity_matrix_to_excel(df, threshold=0.59)
    assert isinstance(excel_bytes, bytes)
    assert len(excel_bytes) > 0

    # Read back generated Excel workbook using openpyxl
    wb = load_workbook(filename=io.BytesIO(excel_bytes))
    ws = wb.active
    assert ws.title == "Similarity Matrix"
    assert ws.cell(row=2, column=2).value == 1.0
